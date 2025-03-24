import os
import logging
import time
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment

from .base_crawler import BaseCrawler, CrawlStats, log_execution_time

# 修改为绝对导入
import sys
import os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config

# 配置日志
logger = logging.getLogger(__name__)

class ExcelUpdater:
    """处理Excel文件更新的类"""
    
    def __init__(self, base_crawler):
        """
        初始化Excel更新器
        
        Args:
            base_crawler: 基础爬虫实例，用于获取WebDriver和其他通用功能
        """
        self.base_crawler = base_crawler
        logger.info("初始化Excel更新器...")
    
    def write_monthly_data(self, worksheet, data, row):
        """
        写入月度数据到Excel
        
        Args:
            worksheet: Excel工作表对象
            data: 包含数据的字典
            row: 要写入的行号
        """
        # 获取工作表的列定义
        sheet_name = worksheet.title
        
        # 导入配置
        # import config
        
        if sheet_name not in config.COLUMN_DEFINITIONS:
            logger.warning(f"工作表 {sheet_name} 没有列定义")
            return
        
        columns = config.COLUMN_DEFINITIONS[sheet_name]
        
        # 写入数据
        for col_idx, col_name in enumerate(columns, 1):
            value = data.get(col_name, '')
            cell = worksheet.cell(row=row, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal='right')
    
    def write_daily_data(self, worksheet, data, last_row, sheet_name):
        """
        写入日频数据到Excel
        
        Args:
            worksheet: Excel工作表对象
            data: 包含数据的列表（通常有多行）
            last_row: 最后一行的行号
            sheet_name: 工作表名称
            
        Returns:
            bool: 是否更新了数据
        """
        # 导入配置
        # import config
        
        if not data:
            return False
        
        # 获取工作表的列定义
        if sheet_name not in config.COLUMN_DEFINITIONS:
            logger.warning(f"工作表 {sheet_name} 没有列定义")
            return False
        
        columns = config.COLUMN_DEFINITIONS[sheet_name]
        
        # 获取最后一行的日期
        last_date = worksheet.cell(row=last_row, column=1).value
        
        # 标记是否有更新
        updated = False
        
        # 遍历数据，从最新的开始
        for idx, row_data in enumerate(data):
            # 获取当前行的日期
            current_date = row_data.get("日期", "")
            if not current_date:
                continue
            
            # 如果当前日期比最后一行的日期更新，则添加新行
            if str(current_date) != str(last_date):
                self.write_single_daily_row(worksheet, row_data, last_row + 1, sheet_name)
                last_row += 1
                updated = True
            else:
                # 如果日期相同，检查是否有需要更新的数据
                # 例如，有些数据可能从"-"更新为具体数值
                need_update = False
                for col_idx, col_name in enumerate(columns, 1):
                    if col_name == '日期':
                        continue
                    
                    # 获取Excel中的当前值
                    current_value = worksheet.cell(row=last_row, column=col_idx).value
                    # 获取新数据中的值
                    new_value = row_data.get(col_name, '')
                    
                    # 检查是否从"-"更新为具体数值
                    if (current_value == '-' or current_value == '') and new_value != '-' and new_value != '':
                        need_update = True
                        break
                
                if need_update:
                    # 更新当前行
                    self.write_single_daily_row(worksheet, row_data, last_row, sheet_name)
                    updated = True
                
                # 已处理最新的一行，不需要继续
                break
        
        return updated
    
    def write_single_daily_row(self, worksheet, row_data, row_num, sheet_name):
        """
        写入单行日频数据
        
        Args:
            worksheet: Excel工作表对象
            row_data: 单行数据字典
            row_num: 要写入的行号
            sheet_name: 工作表名称
        """
        # 导入配置
        # import config
        
        # 获取工作表的列定义
        if sheet_name not in config.COLUMN_DEFINITIONS:
            logger.warning(f"工作表 {sheet_name} 没有列定义")
            return
        
        columns = config.COLUMN_DEFINITIONS[sheet_name]
        
        # 写入数据
        for col_idx, col_name in enumerate(columns, 1):
            value = row_data.get(col_name, '')
            
            # 处理日期格式
            if col_name == '日期' and value:
                try:
                    # 尝试将字符串转换为日期对象
                    if isinstance(value, str):
                        try:
                            value_dt = datetime.strptime(value, '%Y-%m-%d')
                        except ValueError:
                            # 尝试其他日期格式
                            try:
                                value_dt = datetime.strptime(value, '%m/%d/%Y')
                            except ValueError:
                                value_dt = value
                    else:
                        value_dt = value
                    
                    # 格式化日期
                    if isinstance(value_dt, datetime):
                        value = value_dt.strftime('%m/%d/%Y')
                        # 去掉月份和日期的前导零
                        month, day, year = value.split('/')
                        month = month.lstrip('0') if month.startswith('0') and len(month) > 1 else month
                        day = day.lstrip('0') if day.startswith('0') and len(day) > 1 else day
                        value = f"{month}/{day}/{year}"
                except Exception as e:
                    logger.warning(f"日期格式化错误: {str(e)}")
            
            cell = worksheet.cell(row=row_num, column=col_idx, value=value)
            
            # 设置单元格对齐方式
            if sheet_name == 'Shibor':
                cell.alignment = Alignment(horizontal='left')
            elif sheet_name == 'SOFR' and col_idx == 1:
                cell.alignment = Alignment(horizontal='left')
            elif sheet_name == 'SOFR' and col_idx == 2:
                cell.alignment = Alignment(horizontal='left')
            else:
                cell.alignment = Alignment(horizontal='right')
    
    @log_execution_time
    def update_excel(self, results, excel_path):
        """
        更新Excel文件，将爬取的数据写入相应的工作表
        
        Args:
            results: 包含爬取结果的字典，键为工作表名称，值为数据
            excel_path: Excel文件路径
            
        Returns:
            bool: 更新是否成功
        """
        # 导入配置
        # import config
        
        try:
            stats = CrawlStats()  # 创建统计对象
            
            # 如果文件不存在，直接抛出错误
            if not os.path.exists(excel_path):
                raise FileNotFoundError(f"Excel文件不存在: {excel_path}。请确保文件存在于正确的位置。")
            
            logger.info(f"📂 打开Excel文件: {os.path.basename(excel_path)}")
            wb = load_workbook(excel_path)
            
            updated_sheets = []  # 记录已更新的工作表
            
            # 更新各个sheet
            excel_updates = []
            for sheet_name, data in results.items():
                if not data:
                    stats.add_skipped(sheet_name, "数据为空")
                    continue
                
                if sheet_name not in wb.sheetnames:
                    stats.add_skipped(sheet_name, "工作表不存在")
                    logger.warning(f"⚠️ 工作表 {sheet_name} 不存在，跳过更新")
                    continue
                
                ws = wb[sheet_name]
                
                # 查找最后一行数据
                last_row = self.base_crawler.find_last_row(ws)
                
                # 根据数据类型选择不同的处理方法
                if sheet_name in config.MONTHLY_DATA_PAIRS:
                    # 月度数据处理
                    new_date = data.get("日期", "")
                    if not new_date:
                        stats.add_skipped(sheet_name, "数据中缺少日期字段")
                        continue
                    
                    # 获取最后一行的日期值
                    last_date_value = ws.cell(row=last_row, column=1).value
                    
                    # 对Import and Export进行特殊处理
                    if sheet_name == 'Import and Export':
                        # 即使日期相同，也需要检查数据是否从"-"更新为具体数值
                        need_update = False
                        
                        # 如果日期不同，直接更新
                        if str(last_date_value) != str(new_date):
                            need_update = True
                        else:
                            # 日期相同，检查各列数据是否有从"-"更新为具体数值的情况
                            columns = config.COLUMN_DEFINITIONS[sheet_name]
                            for col_idx, col_name in enumerate(columns, 1):
                                if col_name == '日期':
                                    continue
                                
                                # 获取Excel中的当前值
                                current_value = ws.cell(row=last_row, column=col_idx).value
                                # 获取新数据中的值
                                new_value = data.get(col_name, '')
                                
                                # 检查是否从"-"更新为具体数值
                                if (current_value == '-' or current_value == '') and new_value != '-' and new_value != '':
                                    need_update = True
                                    break
                        
                        if need_update:
                            self.write_monthly_data(ws, data, last_row)  # 覆盖当前行
                            excel_updates.append(sheet_name)
                            updated_sheets.append(sheet_name)
                            logger.info(f"📝 更新 {sheet_name}: {new_date}")
                        else:
                            logger.info(f"✓ {sheet_name} 数据已是最新")
                    else:
                        # 其他月度数据的常规处理
                        if str(last_date_value) != str(new_date):
                            self.write_monthly_data(ws, data, last_row + 1)
                            excel_updates.append(sheet_name)
                            updated_sheets.append(sheet_name)
                            logger.info(f"📝 更新 {sheet_name}: {new_date}")
                        else:
                            logger.info(f"✓ {sheet_name} 数据已是最新")
                else:
                    # 日频数据处理（包括汇率数据）
                    update_result = self.write_daily_data(ws, data, last_row, sheet_name)
                    if update_result:
                        excel_updates.append(sheet_name)
                        updated_sheets.append(sheet_name)
                        logger.info(f"📝 更新 {sheet_name}")
            
            # 打印统计摘要
            logger.info("=" * 50)
            stats.print_summary()
            
            # 保存Excel文件
            if excel_updates:
                logger.info(f"💾 保存Excel文件: {os.path.basename(excel_path)}")
                try:
                    wb.save(excel_path)
                    logger.info(f"✅ Excel文件保存成功，已更新 {len(updated_sheets)} 个工作表")
                except Exception as e:
                    logger.error(f"❌ 保存Excel文件时出错: {str(e)}")
                    return False
            else:
                logger.info("ℹ️ 所有工作表数据均已是最新，Excel文件未做修改")
            
            return True
        except Exception as e:
            logger.error(f"❌ 更新Excel过程中出错: {str(e)}", exc_info=True)
            return False
