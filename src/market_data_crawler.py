import requests
import pandas as pd
import logging
from datetime import datetime
import config
from bs4 import BeautifulSoup
import time
import random
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl import Workbook
import zipfile

import platform
from datetime import datetime
import os
import sys
import threading
# 移除signal模块导入，因为它只能在主线程中使用

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.firefox import GeckoDriverManager
from webdriver_manager.microsoft import EdgeChromiumDriverManager
from selenium.common.exceptions import StaleElementReferenceException, TimeoutException, WebDriverException
from selenium.webdriver import ActionChains

import time
import concurrent.futures
from functools import wraps
import fcntl

# 在脚本开头导入并配置连接池
from urllib3.poolmanager import PoolManager
import urllib3

# 禁用所有urllib3警告
urllib3.disable_warnings()

# 增加连接池大小和连接数
class CustomPoolManager(PoolManager):
    def __init__(self, **kwargs):
        kwargs.setdefault('num_pools', 200)
        kwargs.setdefault('maxsize', 200)
        super().__init__(**kwargs)

# 替换默认连接池管理器
urllib3.PoolManager = CustomPoolManager

# 配置日志
logger = logging.getLogger(__name__)

# 添加彩色日志输出
class ColoredFormatter(logging.Formatter):
    """自定义彩色日志格式化器"""

    COLORS = {
        'DEBUG': '\033[94m',     # 蓝色
        'INFO': '\033[92m',      # 绿色
        'WARNING': '\033[93m',   # 黄色
        'ERROR': '\033[91m',     # 红色
        'CRITICAL': '\033[91m\033[1m',  # 红色加粗
        'RESET': '\033[0m'       # 重置颜色
    }

    def format(self, record):
        log_message = super().format(record)
        level_name = record.levelname
        if level_name in self.COLORS:
            return f"{self.COLORS[level_name]}{log_message}{self.COLORS['RESET']}"
        return log_message

def setup_logging(debug=False):
    """设置日志配置"""
    level = logging.DEBUG if debug else logging.INFO

    # 清除现有的处理器
    for handler in logger.handlers[:]:
        logger.removeHandler(handler)

    # 设置日志级别
    logger.setLevel(level)

    # 创建控制台处理器
    console_handler = logging.StreamHandler()
    console_handler.setLevel(level)

    # 创建格式化器
    if os.name == 'posix':  # 在类Unix系统上启用彩色输出
        formatter = ColoredFormatter('%(message)s')
    else:
        formatter = logging.Formatter('%(message)s')

    console_handler.setFormatter(formatter)
    logger.addHandler(console_handler)

    # 文件处理器 - 详细日志保存到文件
    file_handler = logging.FileHandler('market_data_crawler.log')
    file_handler.setLevel(level)

    # 创建logs目录下的日志文件处理器
    # 确保logs目录存在
    logs_dir = 'logs'
    if not os.path.exists(logs_dir):
        os.makedirs(logs_dir)

    # 创建带时间戳的日志文件名
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    logs_file_path = os.path.join(logs_dir, f'crawler_{timestamp}.log')

    # 创建logs目录下的文件处理器
    logs_file_handler = logging.FileHandler(logs_file_path)
    logs_file_handler.setLevel(level)
    file_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    file_handler.setFormatter(file_formatter)
    logger.addHandler(file_handler)

    # 设置logs目录下文件处理器格式
    logs_file_handler.setFormatter(file_formatter)
    logger.addHandler(logs_file_handler)

    # 记录日志启动信息
    logger.info(f"日志已配置：控制台、根目录文件和logs/{os.path.basename(logs_file_path)}")

def log_execution_time(func):
    """记录函数执行时间的装饰器"""
    @wraps(func)
    def wrapper(*args, **kwargs):
        start_time = time.time()
        result = func(*args, **kwargs)
        end_time = time.time()
        elapsed_time = end_time - start_time
        # 只在DEBUG级别记录执行时间，或者在失败时记录
        if result is None:
            logger.warning(f"{func.__name__} 执行失败，耗时: {elapsed_time:.2f} 秒")
        else:
            logger.debug(f"{func.__name__} 执行时间: {elapsed_time:.2f} 秒")
        return result
    return wrapper

def format_error_message(error):
    """格式化错误信息，提取关键部分"""
    error_str = str(error)

    # 如果是Selenium错误，提取主要信息
    if "Session info" in error_str:
        # 提取主要错误信息，去除堆栈跟踪
        main_error = error_str.split('Stacktrace:')[0].strip()
        return main_error

    # 对于其他错误，直接返回错误信息
    return error_str

def log_error(message, error=None, show_traceback=False):
    """统一的错误日志记录函数"""
    if error:
        error_msg = format_error_message(error)
        logger.error(f"{message}: {error_msg}")
        # 只在调试模式下记录完整堆栈
        if show_traceback:
            logger.debug(f"详细错误信息:", exc_info=True)
    else:
        logger.error(message)

def retry_on_timeout(func):
    """重试装饰器，用于处理超时情况"""
    @wraps(func)
    def wrapper(*args, **kwargs):
        max_retries = 3
        retry_count = 0
        while retry_count < max_retries:
            try:
                return func(*args, **kwargs)
            except TimeoutException:
                retry_count += 1
                logger.warning(f"{func.__name__} 第{retry_count}次尝试超时，正在重试...")
                if retry_count >= max_retries:
                    logger.error(f"{func.__name__} 已达到最大重试次数({max_retries})，放弃尝试")
                    return None
                # 每次重试增加等待时间
                time.sleep(2 * retry_count)
            except Exception as e:
                log_error(f"{func.__name__} 发生错误", e, show_traceback=False)
                return None
    return wrapper

# 禁用第三方库的日志
logging.getLogger('urllib3').setLevel(logging.WARNING)
logging.getLogger('selenium').setLevel(logging.WARNING)
logging.getLogger('webdriver_manager').setLevel(logging.WARNING)

# 创建一个统计对象来跟踪成功和失败的爬取
class CrawlStats:
    """爬取统计信息类，用于记录爬取成功、失败和跳过的数据"""

    def __init__(self):
        self.success = []
        self.failure = {}
        self.skipped = {}

    def add_success(self, name):
        self.success.append(name)

    def add_failure(self, name, reason):
        self.failure[name] = reason

    def add_skipped(self, name, reason):
        self.skipped[name] = reason

    def print_summary(self):
        """打印统计摘要并返回摘要文本"""
        # 构建摘要文本
        summary_lines = []
        summary_lines.append("📊 爬取统计摘要")
        summary_lines.append("=" * 50)

        # 成功数据
        if self.success:
            summary_lines.append(f"✅ 成功: {len(self.success)} 项")
            # 每行最多显示4个项目
            for i in range(0, len(self.success), 4):
                chunk = self.success[i:i+4]
                summary_lines.append(f"   {', '.join(chunk)}")

        # 失败数据
        if self.failure:
            summary_lines.append(f"\n❌ 失败: {len(self.failure)} 项")
            for name, reason in self.failure.items():
                summary_lines.append(f"   {name}: {reason}")

        # 跳过数据
        if self.skipped:
            summary_lines.append(f"\n⏭️ 跳过: {len(self.skipped)} 项")
            for name, reason in self.skipped.items():
                summary_lines.append(f"   {name}: {reason}")

        summary_lines.append("=" * 50)

        # 将摘要文本记录到日志
        for line in summary_lines:
            logger.info(line)

        # 返回完整的摘要文本
        return "\n".join(summary_lines)

class MarketDataAnalyzer:
    _driver = None  # 普通WebDriver实例（启用JavaScript）
    _driver_lock = threading.RLock()  # 简单锁，防止并发初始化

    _exchange_rate_driver = None  # 专用于汇率数据的WebDriver实例（禁用JavaScript）
    _exchange_rate_driver_lock = threading.RLock()  # 汇率数据WebDriver的锁

    _daily_driver = None  # 专用于日频数据的WebDriver实例
    _daily_driver_lock = threading.RLock()  # 日频数据WebDriver的锁

    _monthly_driver = None  # 专用于月度数据的WebDriver实例
    _monthly_driver_lock = threading.RLock()  # 月度数据WebDriver的锁

    _instance = None  # 添加单例实例变量

    def __init__(self):
        print("初始化市场数据分析器...")
        # 不再预先初始化WebDriver，而是在需要时按需创建
        # 在多线程环境中不使用信号处理
        # 因为信号处理只能在主线程中使用

        # 单例模式，保存实例引用
        MarketDataAnalyzer._instance = self

    # 移除信号处理方法，因为它只能在主线程中使用

    def _init_driver(self, disable_javascript=False):
        """
        优化的WebDriver初始化方法

        Args:
            disable_javascript: 是否禁用JavaScript，默认为False
        """
        if disable_javascript:
            logger.info("开始初始化WebDriverWithDisableJavascript")
        else:
            logger.info("开始初始化WebDriver")


        import os  # 确保os模块在函数内可用
        system = platform.system()

        # 通用选项
        options = webdriver.ChromeOptions()
        options.add_argument('--headless')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--no-sandbox')
        options.add_argument('--log-level=3')  # 仅显示致命错误
        options.add_argument('--start-maximized')
        options.add_argument('--ignore-certificate-errors')
        options.add_argument('--disable-extensions')
        options.add_argument('--blink-settings=imagesEnabled=false')
        options.add_argument('--disable-blink-features=AutomationControlled')  # 关闭自动化标识
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option('useAutomationExtension', False)
        options.page_load_strategy = 'eager'  # 当DOM就绪时就开始操作，不等待图片等资源

        # 根据参数决定是否禁用JavaScript
        if disable_javascript:
            logger.info("禁用JavaScript模式已启用（用于汇率数据爬取）")
            options.add_experimental_option("prefs", {"profile.managed_default_content_settings.javascript": 2})

        # 添加随机用户代理
        user_agent = self.get_random_user_agent()
        options.add_argument(f'user-agent={user_agent}')
        logger.debug(f"使用用户代理: {user_agent}")

        driver = None
        try:
            # 首先尝试使用Chrome
            from webdriver_manager.chrome import ChromeDriverManager

            driver_dir = ChromeDriverManager().install()
            # 正确的chromedriver路径应该是目录中的chromedriver文件
            driver_path = os.path.join(os.path.dirname(driver_dir), 'chromedriver')

            # 确保文件有执行权限
            os.chmod(driver_path, 0o755)

            # driver_path = ChromeDriverManager().install()
            # driver_path ='/Users/dieselchen/.wdm/drivers/chromedriver/mac64/134.0.6998.165/chromedriver-mac-x64/chromedriver'
            # driver_path='/root/.wdm/drivers/chromedriver/linux64/140.0.7339.80/chromedriver-linux64/chromedriver'

            service = Service(executable_path=driver_path)

            # 创建driver并修改navigator.webdriver
            driver = webdriver.Chrome(service=service, options=options)

            logger.info("成功初始化 Chrome WebDriver")
        except Exception as e:
            logger.warning(f"Chrome WebDriver 初始化失败: {str(e)}")

            try:
                # 尝试使用Edge
                from webdriver_manager.microsoft import EdgeChromiumDriverManager

                edge_options = webdriver.EdgeOptions()
                for arg in options.arguments:
                    edge_options.add_argument(arg)
                edge_options.use_chromium = True
                edge_options.page_load_strategy = 'eager'

                driver_path = EdgeChromiumDriverManager().install()

                # 创建一个空的日志文件对象来抑制输出
                if system == "Windows":
                    null_output = open(os.devnull, 'w')
                    service = Service(executable_path=driver_path, log_output=null_output)
                else:
                    service = Service(executable_path=driver_path)

                driver = webdriver.Edge(service=service, options=edge_options)

                # 执行JavaScript修改webdriver标识
                driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {
                    'source': '''
                        Object.defineProperty(navigator, 'webdriver', {
                            get: () => undefined
                        });
                        // 修改语言标识以增加随机性
                        Object.defineProperty(navigator, 'languages', {
                            get: () => ['zh-CN', 'zh', 'en-US', 'en']
                        });
                    '''
                })

                logger.info("成功初始化 Edge WebDriver")
            except Exception as e:
                logger.warning(f"Edge WebDriver 初始化失败: {str(e)}")

                try:
                    # 最后尝试Firefox
                    from webdriver_manager.firefox import GeckoDriverManager

                    firefox_options = webdriver.FirefoxOptions()
                    for arg in options.arguments:
                        if not arg.startswith('--disable-dev-shm-usage') and not arg.startswith('--no-sandbox'):
                            firefox_options.add_argument(arg)
                    firefox_options.page_load_strategy = 'eager'
                    firefox_options.add_argument('--log-level=3')  # 仅显示致命错误

                    # Firefox特有的性能设置
                    firefox_profile = webdriver.FirefoxProfile()
                    firefox_profile.set_preference("dom.webdriver.enabled", False)
                    firefox_profile.set_preference('useAutomationExtension', False)
                    firefox_profile.set_preference("general.useragent.override", user_agent)
                    firefox_profile.update_preferences()
                    firefox_options.profile = firefox_profile

                    driver_path = GeckoDriverManager().install()

                    service = Service(executable_path=driver_path)
                    driver = webdriver.Firefox(service=service, options=firefox_options)
                    logger.info("成功初始化 Firefox WebDriver")
                except Exception as e:
                    logger.error(f"所有WebDriver初始化失败: {str(e)}")
                    raise

        return driver

    def get_driver(self, driver_type='default'):
        """
        获取WebDriver实例，如果不存在则初始化

        Args:
            driver_type: WebDriver类型，可选值：'default'(默认), 'exchange_rate'(汇率数据), 'daily'(日频数据), 'monthly'(月度数据)

        Returns:
            WebDriver实例
        """
        if driver_type == 'exchange_rate':
            # 获取专用于汇率数据的WebDriver实例（禁用JavaScript）
            with self._exchange_rate_driver_lock:
                if self._exchange_rate_driver is None:
                    self._exchange_rate_driver = self._init_driver(disable_javascript=True)
                return self._exchange_rate_driver
        elif driver_type == 'daily':
            # 获取专用于日频数据的WebDriver实例
            with self._daily_driver_lock:
                if self._daily_driver is None:
                    self._daily_driver = self._init_driver(disable_javascript=False)
                return self._daily_driver
        elif driver_type == 'monthly':
            # 获取专用于月度数据的WebDriver实例
            with self._monthly_driver_lock:
                if self._monthly_driver is None:
                    self._monthly_driver = self._init_driver(disable_javascript=False)
                return self._monthly_driver
        else:
            # 获取普通WebDriver实例（启用JavaScript）
            with self._driver_lock:
                if self._driver is None:
                    self._driver = self._init_driver(disable_javascript=False)
                return self._driver

    def close_driver(self, driver_type='default'):
        """
        关闭WebDriver实例

        Args:
            driver_type: WebDriver类型，可选值：'default'(默认), 'exchange_rate'(汇率数据), 'daily'(日频数据), 'monthly'(月度数据)
        """
        if driver_type == 'exchange_rate':
            # 关闭汇率数据专用的WebDriver实例
            with self._exchange_rate_driver_lock:
                if self._exchange_rate_driver:
                    try:
                        self._exchange_rate_driver.quit()
                        logger.info("汇率数据WebDriver已关闭")
                    except Exception as e:
                        logger.warning(f"关闭汇率数据WebDriver时出错: {str(e)}")
                    finally:
                        self._exchange_rate_driver = None
        elif driver_type == 'daily':
            # 关闭日频数据专用的WebDriver实例
            with self._daily_driver_lock:
                if self._daily_driver:
                    try:
                        self._daily_driver.quit()
                        logger.info("日频数据WebDriver已关闭")
                    except Exception as e:
                        logger.warning(f"关闭日频数据WebDriver时出错: {str(e)}")
                    finally:
                        self._daily_driver = None
        elif driver_type == 'monthly':
            # 关闭月度数据专用的WebDriver实例
            with self._monthly_driver_lock:
                if self._monthly_driver:
                    try:
                        self._monthly_driver.quit()
                        logger.info("月度数据WebDriver已关闭")
                    except Exception as e:
                        logger.warning(f"关闭月度数据WebDriver时出错: {str(e)}")
                    finally:
                        self._monthly_driver = None
        else:
            # 关闭普通WebDriver实例
            with self._driver_lock:
                if self._driver:
                    try:
                        self._driver.quit()
                        logger.info("普通WebDriver已关闭")
                    except Exception as e:
                        logger.warning(f"关闭普通WebDriver时出错: {str(e)}")
                    finally:
                        self._driver = None

    def get_random_user_agent(self):
        user_agents = [
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/93.0.4577.82 Safari/537.36",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/94.0.4606.81 Safari/537.36",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/94.0.4606.81 Safari/537.36",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:93.0) Gecko/20100101 Firefox/93.0",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:93.0) Gecko/20100101 Firefox/93.0"
        ]
        return random.choice(user_agents)


    def format_exchange_rate_date(self,raw_date):
        # 解析中文月份
        dt = datetime.strptime(raw_date, "%m月 %d, %Y")

        # 判断操作系统
        if platform.system() == "Windows":
            return dt.strftime("%Y/%#m/%d")
        else:  # Linux/macOS
            return dt.strftime("%Y/%-m/%d")

    def format_stee_price_date(self,raw_date):
        # 解析中文月份
        dt = datetime.strptime(raw_date, "%Y/%m/%d")

        # 判断操作系统
        if platform.system() == "Windows":
            return dt.strftime("%Y/%#m/%d")
        else:  # Linux/macOS
            return dt.strftime("%Y/%-m/%d")

    def format_shibor_rate_date(self,raw_date):
        # 解析中文月份
        dt = datetime.strptime(raw_date, "%Y-%m-%d")

        # 判断操作系统
        if platform.system() == "Windows":
            return dt.strftime("%Y/%#m/%d")
        else:  # Linux/macOS
            return dt.strftime("%Y/%-m/%d")

    def format_sofr_date(self, raw_date):
        # 获取当前年份
        current_year = datetime.now().year
        # 拼接年份、月份和日期
        full_date_str = f"{current_year}/{raw_date}"

        try:
            # 解析日期字符串为 datetime 对象
            dt = datetime.strptime(full_date_str, "%Y/%m/%d")
            # 判断操作系统
            if platform.system() == "Windows":
                return dt.strftime("%Y/%#m/%d")
            else:  # Linux/macOS
                return dt.strftime("%Y/%-m/%d")
        except ValueError:
            print(f"日期解析失败，输入的日期 {raw_date} 格式可能不正确。")
            return None

    def format_ester_date(self, raw_date):
        # 解析中文月份
        dt = datetime.strptime(raw_date, "%m/%d/%Y")

        # 判断操作系统
        if platform.system() == "Windows":
            return dt.strftime("%Y/%#m/%d")
        else:  # Linux/macOS
            return dt.strftime("%Y/%-m/%d")

    def format_jpy_rate_date(self, raw_date):
        # 解析中文月份
        dt = datetime.strptime(raw_date, "%m-%d-%Y")

        # 判断操作系统
        if platform.system() == "Windows":
            return dt.strftime("%Y/%#m/%d")
        else:  # Linux/macOS
            return dt.strftime("%Y/%-m/%d")

    def format_lpr_date(self, raw_date):
        # 解析中文月份
        dt = datetime.strptime(raw_date, "%Y-%m-%d")

        # 判断操作系统
        if platform.system() == "Windows":
            return dt.strftime("%Y/%#m/%d")
        else:  # Linux/macOS
            return dt.strftime("%Y/%-m/%d")

    def format_us_interest_rate_date(self, raw_date):
        # 解析中文月份
        dt = datetime.strptime(raw_date, "%Y-%m-%d")

        # 判断操作系统
        if platform.system() == "Windows":
            return dt.strftime("%Y/%#m/%d")
        else:  # Linux/macOS
            return dt.strftime("%Y/%-m/%d")

    @log_execution_time
    @retry_on_timeout
    def crawl_exchange_rate(self, url):
        """优化后的汇率数据爬取方法（带详细调试日志）"""
        # 获取专用于汇率数据的WebDriver实例（禁用JavaScript）
        driver = self.get_driver(driver_type='exchange_rate')
        logger.info(f"开始爬取汇率数据：{url}")

        try:

            # 设置超时策略
            driver.set_page_load_timeout(10)
            wait = WebDriverWait(driver, 10, poll_frequency=0.25)

            try:
                logger.debug("尝试加载页面...")
                driver.get(url)
            except TimeoutException:
                logger.warning("页面加载超时，强制停止")
                driver.execute_script("window.stop();")

            # 表格定位策略优化
            try:
                logger.debug("定位数据表格...")
                table = wait.until(EC.presence_of_element_located((
                    By.CSS_SELECTOR, 'table.freeze-column-w-1'
                )))
                logger.debug("表格定位成功")
            except TimeoutException as e:
                logger.error("❌ 表格定位失败，可能原因：")
                logger.error("1. 页面结构已变更")
                logger.error("2. 反爬机制触发")
                logger.error("3. 网络请求被拦截")
                raise

            # 数据行获取策略
            def _load_rows(driver):
                """带滚动加载的行获取函数"""
                last_height = driver.execute_script("return document.body.scrollHeight")
                for _ in range(3):  # 最多滚动3次
                    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    new_height = driver.execute_script("return document.body.scrollHeight")
                    if new_height == last_height:
                        break
                    last_height = new_height

                rows = driver.find_elements(
                    By.CSS_SELECTOR,
                    "tr.historical-data-v2_price__atUfP:not(:empty)"
                )
                return rows if len(rows) > 5 else None  # 至少需要5行数据

            try:
                logger.debug("尝试获取数据行...")
                rows = wait.until(
                    lambda d: _load_rows(d) or (_load_rows(d) and False),
                    message="数据行加载失败"
                )
            except TimeoutException:
                logger.error("数据行加载超时，可能原因：")
                logger.error("1. 滚动加载未触发")
                logger.error("2. 反爬验证未通过")
                return None

            # 数据解析优化
            results = []
            required_columns = {"收盘", "开盘", "高", "低"}
            for idx, row in enumerate(rows[:10]):  # 限制处理前100行
                try:


                    # 动态定位元素
                    date_cell = row.find_element(By.CSS_SELECTOR, "td:first-child time")
                    cells = row.find_elements(By.CSS_SELECTOR, "td:not([style*='display:none'])")


                    # 数据校验
                    if len(cells) < 5:
                        logger.debug(f"跳过第 {idx} 行，数据列不足")
                        continue

                    # 构建数据记录
                    record = {
                        "日期": self.format_exchange_rate_date(date_cell.text),
                        "收盘": cells[1].text.strip(),
                        "开盘": cells[2].text.strip(),
                        "高": cells[3].text.strip(),
                        "低": cells[4].text.strip()
                    }

                    if url == 'https://cn.investing.com/rates-bonds/u.s.-10-year-bond-yield-historical-data':
                        record["涨跌幅"] = cells[5].text.strip()
                    else:
                        record["涨跌幅"] = cells[6].text.strip()


                    results.append(record)

                except StaleElementReferenceException:
                    logger.debug(f"第 {idx} 行元素失效，重新获取中...")
                    rows = driver.find_elements(
                        By.CSS_SELECTOR,
                        "tr.historical-data-v2_price__atUfP:not(:empty)"
                    )
                    continue
                except Exception as e:
                    logger.debug(f"第 {idx} 行解析异常：{str(e)}")
                    continue

            logger.debug(f"成功解析 {len(results)} 条有效记录")
            return results

        except Exception as e:
            logger.error(f"爬取过程异常：{str(e)}")
            logger.debug(f"异常堆栈：", exc_info=True)
            return None


    def find_last_row(self, sheet):
        """
        改进的查找最后一行方法：逆向查找第一个非空行
        """
        for row in reversed(range(1, sheet.max_row + 1)):
            if any(cell.value for cell in sheet[row]):
                return row
        return 1  # 如果全为空，从第一行开始

    def write_monthly_data(self, worksheet, data, row):
        """
        写入月度数据到Excel

        Args:
            worksheet: Excel工作表对象
            data: 包含数据的字典
            row: 要写入的行号
        """
        # 获取工作表名称
        sheet_name = worksheet.title

        # 获取该工作表对应的列定义
        if sheet_name in config.COLUMN_DEFINITIONS:
            columns = config.COLUMN_DEFINITIONS[sheet_name]
        else:
            logger.warning(f"未找到 {sheet_name} 的列定义，使用默认列")
            columns = ['日期']

        # 写入数据
        for col_idx, col_name in enumerate(columns, 1):
            value = data.get(col_name, '')
            cell = worksheet.cell(row=row, column=col_idx, value=value)
            if sheet_name == 'Import and Export' and col_idx == 1:
                cell.alignment = Alignment(horizontal='left')
            elif sheet_name == 'Money Supply' and col_idx == 1:
                cell.alignment = Alignment(horizontal='left')
            elif sheet_name == 'PPI' and col_idx == 1:
                cell.alignment = Alignment(horizontal='left')
            elif sheet_name == 'CPI' and col_idx == 1:
                cell.alignment = Alignment(horizontal='left')
            elif sheet_name == 'PMI' and col_idx == 1:
                cell.alignment = Alignment(horizontal='left')
            elif sheet_name == 'New Bank Loan Addition' and col_idx == 1:
                cell.alignment = Alignment(horizontal='left')
            elif sheet_name == 'US Interest Rate' and col_idx == 1:
                cell.alignment = Alignment(horizontal='left')
            else:
                cell.alignment = Alignment(horizontal='right')

        logger.info(f"已在 {sheet_name} 的第 {row} 行写入月度数据")

    def write_daily_data(self, worksheet, data, last_row, sheet_name):
        """
        写入日频数据到Excel

        Args:
            worksheet: Excel工作表对象
            data: 包含数据的列表（通常有多行）
            last_row: 最后一行的行号
            sheet_name: 工作表名称

        Returns:
            bool: 是否更新了数据
        """
        # 检查数据是否为空
        if not data or len(data) < 1:
            logger.error(f"{sheet_name}: 数据不足，无法写入")
            return False

        # 获取最新数据的日期
        new_date_str = data[0].get("日期", "")
        if not new_date_str:
            logger.error(f"{sheet_name}: 数据中缺少日期字段")
            return False

        # 获取最后一行的日期值
        last_date_value = worksheet.cell(row=last_row, column=1).value

        # 解析现有日期和新日期为datetime对象，用于比较
        new_date_obj = None
        last_date_obj = None

        # 解析新日期为datetime对象
        try:
            if '/' in new_date_str:
                year, month, day = map(int, new_date_str.split('/'))
                new_date_obj = datetime(year, month, day)
            elif '-' in new_date_str:
                year, month, day = map(int, new_date_str.split('-'))
                new_date_obj = datetime(year, month, day)
        except Exception as e:
            logger.warning(f"{sheet_name}: 解析新日期 '{new_date_str}' 失败: {str(e)}")


        # 解析现有日期
        if isinstance(last_date_value, datetime):
            last_date_obj = last_date_value
        else:
            try:
                if last_date_value:
                    if sheet_name == 'SOFR':
                        month, day, year = map(int, str(last_date_value).split('/'))
                        last_date_obj = datetime(year, month, day)
                    elif sheet_name == 'Shibor':
                        year, month, day = map(int, str(last_date_value).split('-'))
                        last_date_obj = datetime(year, month, day)
                    else:
                        year, month, day = map(int, str(last_date_value).split('/'))
                        last_date_obj = datetime(year, month, day)
            except Exception as e:
                logger.warning(
                    f"{sheet_name}: 解析最后一行日期 '{last_date_value}' 失败: {str(e)}"
                    f"last_date_value 的值是: {last_date_value}，类型是: {type(last_date_value)} "
                )

        if new_date_obj is None or last_date_obj is None:
            # 若有日期对象为 None，则记录警告信息
            logger.warning(
                f"{sheet_name}: 日期对象比较失败，请重试后不行联系管理员。"
                f"last_date_value 的值是: {last_date_value}，类型是: {type(last_date_value)} "
                f"new_date_str 的值是: {new_date_str}，类型是: {type(new_date_str)}"
            )
        # 若两个日期对象都不为 None，则比较日期
        elif new_date_obj.date() == last_date_obj.date():
            # 若日期相同，则记录调试信息并返回 False
            logger.debug(
                f"{sheet_name}: 日期对象比较相同 ({new_date_obj.date()} == {last_date_obj.date()})，数据已是最新，无需更新"
            )
            return False

        # 在数据列表中查找最后一行日期的位置
        last_date_index = -1

        # 使用datetime对象比较查找
        if last_date_obj:
            for i, item in enumerate(data):
                item_date_str = item.get("日期", "")
                try:
                    if '/' in item_date_str:
                        year, month, day = map(int, item_date_str.split('/'))
                        item_date = datetime(year, month, day)
                    elif '-' in item_date_str:
                        year, month, day = map(int, item_date_str.split('-'))
                        item_date = datetime(year, month, day)
                    else:
                        continue

                    if item_date.date() == last_date_obj.date():
                        logger.debug(f"{sheet_name}: 找到最后一行日期(对象比较): {item_date} 在索引 {i} 即将插入{i}个新数据 刷新最后一行数据")
                        last_date_index = i
                        break
                except Exception as e:
                    logger.debug(f"{sheet_name}: 解析日期 '{item_date_str}' 失败: {str(e)}")
                    continue

        # 如果找到了最后一行日期
        if last_date_index != -1:
            # 用找到的数据覆盖最后一行
            self.write_single_daily_row(worksheet, data[last_date_index], last_row, sheet_name)
            logger.debug(f"{sheet_name}: 已更新第 {last_row} 行数据")

            # 将最后一行日期之前的数据倒序插入
            for i in range(last_date_index - 1, -1, -1):
                # 插入新行
                target_row = last_row + (last_date_index - i)
                self.write_single_daily_row(worksheet, data[i], target_row, sheet_name)
                logger.debug(f"{sheet_name}: 已在第 {target_row} 行插入新数据")

            return True
        else:
            # 如果没找到最后一行日期，记录日志
            logger.warning(f"{sheet_name}: 爬取的最新数据并没有匹配上现有数据，无法更新.现有数据{data}，最后一行日期{last_date_obj}")
            # 进一步处理：视为长期未更新，将现有数据倒序追加到Excel
            try:
                logger.warning(f"{sheet_name}: 未找到匹配日期，判定为长期未更新。将倒序追加 {len(data)} 条数据到Excel")
                # 从最旧到最新写入：因为data[0]通常是最新，因此倒序遍历
                for idx in range(len(data) - 1, -1, -1):
                    target_row = last_row + (len(data) - idx)
                    self.write_single_daily_row(worksheet, data[idx], target_row, sheet_name)
                    logger.debug(f"{sheet_name}: 倒序追加 —— 已在第 {target_row} 行写入索引 {idx} 的数据")
                return True
            except Exception as e:
                logger.error(f"{sheet_name}: 倒序追加写入失败: {str(e)}")
                return False

    def write_single_daily_row(self, worksheet, row_data, row_num, sheet_name):
        """
        写入单行日频数据

        Args:
            worksheet: Excel工作表对象
            row_data: 单行数据字典
            row_num: 要写入的行号
            sheet_name: 工作表名称
        """
        # 获取该工作表对应的列定义
        if sheet_name in config.COLUMN_DEFINITIONS:
            columns = config.COLUMN_DEFINITIONS[sheet_name]
        elif sheet_name in config.CURRENCY_PAIRS:
            # 汇率数据使用通用列定义
            if sheet_name == 'USD 10Y':
                columns = config.COLUMN_DEFINITIONS['USD 10Y']
            else:
                columns = config.COLUMN_DEFINITIONS['CURRENCY']
        else:
            logger.warning(f"未找到 {sheet_name} 的列定义，使用默认列")
            columns = ['日期']

        # 写入数据
        for col_idx, col_name in enumerate(columns, 1):
            value = row_data.get(col_name, '')
            if sheet_name == 'Shibor' and col_idx == 1:
                value_dt = datetime.strptime(value, '%Y/%m/%d')
                if isinstance(value_dt, datetime):
                    value = value_dt.strftime('%Y-%m-%d')
            if sheet_name == 'SOFR' and col_idx == 1:
                value_dt = datetime.strptime(value, '%Y/%m/%d')
                if isinstance(value_dt, datetime):
                    value = value_dt.strftime('%m/%d/%Y')
                    # 去掉月份和日期的前导零
                    month, day, year = value.split('/')
                    month = month.lstrip('0') if month.startswith('0') and len(month) > 1 else month
                    day = day.lstrip('0') if day.startswith('0') and len(day) > 1 else day
                    value = f"{month}/{day}/{year}"
            cell = worksheet.cell(row=row_num, column=col_idx, value=value)
            if sheet_name == 'Shibor':
                cell.alignment = Alignment(horizontal='left')
            elif sheet_name == 'SOFR' and col_idx == 1:
                cell.alignment = Alignment(horizontal='left')
            elif sheet_name == 'SOFR' and col_idx == 2:
                cell.alignment = Alignment(horizontal='left')
            else:
                cell.alignment = Alignment(horizontal='right')

    def update_excel(self):
        """
        更新现有Excel文件，追加数据到对应sheet的最后一行（并发执行版本）
        汇率数据、日频数据和月度数据分别使用不同的WebDriver实例并行爬取
        """
        stats = CrawlStats()  # 创建统计对象

        try:
            results = {}

            # 创建进度跟踪器
            total_tasks = len(config.CURRENCY_PAIRS) + len(config.DAILY_DATA_PAIRS) + len(config.MONTHLY_DATA_PAIRS)
            completed_tasks = 0

            # 打印任务总览
            logger.info("=" * 50)
            logger.info("🚀 开始数据爬取任务（三线程并发执行模式）")
            logger.info("=" * 50)
            logger.info(f"📊 汇率数据: {len(config.CURRENCY_PAIRS)} 项")
            logger.info(f"📈 日频数据: {len(config.DAILY_DATA_PAIRS)} 项")
            logger.info(f"📅 月度数据: {len(config.MONTHLY_DATA_PAIRS)} 项")
            logger.info(f"🔄 总任务数: {total_tasks} 项")
            logger.info("=" * 50)

            # 初始化三个WebDriver实例
            logger.info("⚙️ 初始化三个WebDriver实例...")
            # WebDriver实例将通过get_driver()方法按需初始化

            # 更新进度的辅助函数
            def update_progress(sheet_name, data_type, success=True, error_msg=None):
                nonlocal completed_tasks
                completed_tasks += 1
                progress = int(completed_tasks / total_tasks * 100)
                progress_bar = "█" * (progress // 5) + "░" * (20 - progress // 5)

                if success:
                    logger.info(f"✅ [{progress:3d}%] |{progress_bar}| {sheet_name} ({data_type})")
                elif error_msg:
                    logger.error(f"❌ [{progress:3d}%] |{progress_bar}| {sheet_name} ({data_type}): {error_msg}")
                else:
                    logger.warning(f"⚠️ [{progress:3d}%] |{progress_bar}| {sheet_name} ({data_type}): 数据为空")

            # 定义爬取函数
            def crawl_exchange_rate_task(pair, url):
                try:
                    # 使用专用于汇率数据的WebDriver
                    driver = self.get_driver(driver_type='exchange_rate')
                    data = self.crawl_exchange_rate(url)
                    if data:
                        with results_lock:
                            results[pair] = data
                        stats.add_success(pair)
                        update_progress(pair, "currency")
                        return True
                    else:
                        stats.add_failure(pair, "爬取返回空数据")
                        update_progress(pair, "currency", False)
                        return False
                except Exception as e:
                    stats.add_failure(pair, str(e))
                    update_progress(pair, "currency", False, str(e))
                    return False

            def crawl_daily_task(sheet_name, info):
                try:
                    # 使用专用于日频数据的WebDriver
                    driver = self.get_driver(driver_type='daily')
                    crawler_method = getattr(self, info['crawler'])
                    data = crawler_method(info['url'])

                    if data:
                        with results_lock:
                            results[sheet_name] = data
                        stats.add_success(sheet_name)
                        update_progress(sheet_name, "daily")
                        return True
                    else:
                        stats.add_failure(sheet_name, "爬取返回空数据")
                        update_progress(sheet_name, "daily", False)
                        return False
                except Exception as e:
                    stats.add_failure(sheet_name, str(e))
                    update_progress(sheet_name, "daily", False, str(e))
                    return False

            def crawl_monthly_task(sheet_name, info):
                try:
                    # 使用专用于月度数据的WebDriver
                    driver = self.get_driver(driver_type='monthly')
                    crawler_method = getattr(self, info['crawler'])
                    data = crawler_method(info['url'])

                    if data:
                        # 对于月度数据，只保留第一行
                        if isinstance(data, list) and len(data) > 0:
                            with results_lock:
                                results[sheet_name] = data[0]
                        else:
                            with results_lock:
                                results[sheet_name] = data
                        stats.add_success(sheet_name)
                        update_progress(sheet_name, "monthly")
                        return True
                    else:
                        stats.add_failure(sheet_name, "爬取返回空数据")
                        update_progress(sheet_name, "monthly", False)
                        return False
                except Exception as e:
                    stats.add_failure(sheet_name, str(e))
                    update_progress(sheet_name, "monthly", False, str(e))
                    return False

            # 使用线程锁保护共享资源
            results_lock = threading.RLock()

            # 创建三个线程池，分别用于汇率数据、日频数据和月度数据
            with concurrent.futures.ThreadPoolExecutor(max_workers=1) as exchange_rate_executor, \
                 concurrent.futures.ThreadPoolExecutor(max_workers=1) as daily_executor, \
                 concurrent.futures.ThreadPoolExecutor(max_workers=1) as monthly_executor:

                # 1. 提交汇率数据爬取任务
                logger.info("开始爬取汇率数据（并发执行）...")
                exchange_rate_futures = []
                for pair, url in config.CURRENCY_PAIRS.items():
                    future = exchange_rate_executor.submit(crawl_exchange_rate_task, pair, url)
                    exchange_rate_futures.append(future)

                # 2. 提交日频数据爬取任务
                logger.info("开始爬取日频数据（并发执行）...")
                daily_futures = []
                for sheet_name, info in config.DAILY_DATA_PAIRS.items():
                    future = daily_executor.submit(crawl_daily_task, sheet_name, info)
                    daily_futures.append(future)

                # 3. 提交月度数据爬取任务
                logger.info("开始爬取月度数据（并发执行）...")
                monthly_futures = []
                for sheet_name, info in config.MONTHLY_DATA_PAIRS.items():
                    future = monthly_executor.submit(crawl_monthly_task, sheet_name, info)
                    monthly_futures.append(future)

                # 等待所有任务完成
                logger.info("等待所有爬取任务完成...")
                concurrent.futures.wait(exchange_rate_futures + daily_futures + monthly_futures)

            # 关闭WebDriver实例
            logger.info("爬取任务完成，关闭WebDriver实例...")
            self.close_driver(driver_type='exchange_rate')  # 关闭汇率数据WebDriver
            self.close_driver(driver_type='daily')         # 关闭日频数据WebDriver
            self.close_driver(driver_type='monthly')       # 关闭月度数据WebDriver

            logger.info("=" * 50)
            logger.info("🏁 数据爬取完成，准备更新Excel文件...")

            # 4. 更新Excel文件
            logger.info("开始更新Excel文件...")
            excel_path = config.EXCEL_OUTPUT_PATH
            logger.info(f"📂 打开Excel文件: {os.path.basename(excel_path)}")

            # 如果文件不存在，直接抛出错误
            if not os.path.exists(excel_path):
                raise FileNotFoundError(f"Excel文件不存在: {excel_path}。请确保文件存在于正确的位置。")

            # 跨进程文件锁，防止并发读写导致损坏
            lock_path = excel_path + ".lock"
            lock_fd = None
            try:
                lock_fd = open(lock_path, 'w')
                fcntl.flock(lock_fd, fcntl.LOCK_EX)
                logger.debug("已获取Excel文件锁")

                # 尝试加载工作簿；若失败则不修改原文件，直接返回失败
                try:
                    wb = load_workbook(excel_path)
                except Exception as e:
                    logger.error(f"无法打开Excel文件（可能不是有效的xlsx或被占用）：{str(e)}")
                    return False
            except Exception as le:
                logger.error(f"获取Excel文件锁失败: {str(le)}")
                return False

            updated_sheets = []  # 记录已更新的工作表

            # 更新各个sheet
            excel_updates = []
            for sheet_name, data in results.items():
                if not data:
                    stats.add_skipped(sheet_name, "数据为空")
                    continue

                if sheet_name not in wb.sheetnames:
                    stats.add_skipped(sheet_name, "工作表不存在")
                    logger.warning(f"⚠️ 工作表 {sheet_name} 不存在，跳过更新")
                    continue

                ws = wb[sheet_name]

                # 查找最后一行数据
                last_row = self.find_last_row(ws)


                # 根据数据类型选择不同的处理方法
                if sheet_name in config.MONTHLY_DATA_PAIRS:
                    # 月度数据处理
                    new_date = data.get("日期", "")
                    if not new_date:
                        stats.add_skipped(sheet_name, "数据中缺少日期字段")
                        continue

                    # 获取最后一行的日期值
                    last_date_value = ws.cell(row=last_row, column=1).value

                    # 对Import and Export进行特殊处理
                    if sheet_name == 'Import and Export':
                        # 检查最后一行的数据是否完整（没有"-"）
                        last_row_complete = True
                        columns = config.COLUMN_DEFINITIONS[sheet_name]

                        # 检查最后一行的每个单元格（除了日期列）
                        for col_idx, col_name in enumerate(columns, 1):
                            if col_name == '日期':
                                continue

                            current_value = ws.cell(row=last_row, column=col_idx).value
                            if current_value == '-' or current_value == '':
                                last_row_complete = False
                                break

                        if not last_row_complete:
                            # 如果最后一行不完整，用新数据更新这一行
                            self.write_monthly_data(ws, data, last_row)
                            excel_updates.append(sheet_name)
                            updated_sheets.append(sheet_name)
                            logger.info(f"📝 更新不完整行 {sheet_name}: {new_date}")
                        elif str(last_date_value) != str(new_date):
                            # 如果最后一行完整且日期不同，写入新行
                            self.write_monthly_data(ws, data, last_row + 1)
                            excel_updates.append(sheet_name)
                            updated_sheets.append(sheet_name)
                            logger.info(f"📝 添加新行 {sheet_name}: {new_date}")
                        else:
                            logger.info(f"✓ {sheet_name} 数据已是最新且完整")
                    else:
                        # 其他月度数据的常规处理
                        if str(last_date_value) != str(new_date):
                            self.write_monthly_data(ws, data, last_row + 1)
                            excel_updates.append(sheet_name)
                            updated_sheets.append(sheet_name)
                            logger.info(f"📝 更新 {sheet_name}: {new_date}")
                        else:
                            logger.info(f"✓ {sheet_name} 数据已是最新")
                else:
                    # 日频数据处理（包括汇率数据）
                    update_result = self.write_daily_data(ws, data, last_row, sheet_name)
                    if update_result:
                        excel_updates.append(sheet_name)
                        updated_sheets.append(sheet_name)
                        logger.info(f"📝 更新 {sheet_name}")

            # 打印统计摘要并获取摘要文本
            logger.info("=" * 50)
            summary_text = stats.print_summary()

            # 添加一个特殊的日志消息，标记为摘要信息（前端据此收集并在收到 SHOW_SUMMARY 时显示）
            try:
                logger.info("SUMMARY_START")
                for line in summary_text.splitlines():
                    if line.strip():
                        logger.info(line)
                logger.info("SUMMARY_END")
            except Exception:
                # 回退：直接输出文本
                logger.info(summary_text)

            # 保存Excel文件
            if excel_updates:
                logger.info(f"💾 保存Excel文件: {os.path.basename(excel_path)}")
                try:
                    tmp_path = excel_path + ".tmp"
                    wb.save(tmp_path)
                    os.replace(tmp_path, excel_path)
                    logger.info(f"✅ Excel文件保存成功，已更新 {len(updated_sheets)} 个工作表")
                except Exception as e:
                    logger.error(f"❌ 保存Excel文件时出错: {str(e)}")
                    return False
            else:
                logger.info("ℹ️ 所有工作表数据均已是最新，Excel文件未做修改")

            return results
        except Exception as e:
            logger.error(f"❌ 更新Excel过程中出错: {str(e)}", exc_info=True)
            return False
        finally:
            # 释放文件锁
            try:
                if lock_fd is not None:
                    fcntl.flock(lock_fd, fcntl.LOCK_UN)
                    lock_fd.close()
                    logger.debug("已释放Excel文件锁")
                    # 显式标记Excel已释放，供前端/后端完成判定
                    logger.info("EXCEL_UNLOCKED")
            except Exception:
                pass

    @log_execution_time
    @retry_on_timeout
    def crawl_steel_price(self, url):
        """
        爬取钢铁价格数据（优化版）
        """
        driver = self.get_driver(driver_type='daily')
        logger.debug(f"正在请求URL: {url}")

        try:

            driver.set_page_load_timeout(30)
            wait = WebDriverWait(driver, 20, poll_frequency=0.25)

            driver.get(url)

            # 使用显式等待，减少固定等待时间
            wait.until(EC.element_to_be_clickable((By.XPATH, '//span[text()="相对价格指数走势图"]'))).click()

            # 等待数据完全加载
            wait.until(EC.presence_of_element_located((By.XPATH, '//td[contains(text(),"/") and string-length(text())>8]')))

            # 获取表格引用
            table = driver.find_element(By.XPATH, '//table[contains(@class,"detailTab")]')

            # 单次获取所有需要的数据 - 修改为获取前10行
            rows = table.find_elements(By.XPATH, './/tbody/tr[position()<=10]')
            data = []

            for row in rows:
                try:
                    # 实时获取当前行元素
                    cells = row.find_elements(By.XPATH, './/td[not(contains(@style,"none"))]')


                    # 数据校验
                    if len(cells) < 10:
                        logger.debug(f"Steel price: 跳过无效行，列数：{len(cells)}")
                        continue

                    # 立即提取文本内容
                    cell_texts = [cell.text for cell in cells]


                    # 动态映射字段
                    item = {
                        "日期": self.format_stee_price_date(cells[0].get_attribute('textContent').strip()),
                        "本日": cells[1].text.strip(),
                        "昨日": cells[2].text.strip(),
                        "日环比": cells[3].text.strip(),
                        "上周": cells[4].text.strip(),
                        "周环比": cells[5].text.strip(),
                        "上月度": cells[6].text.strip(),
                        "与上月比": cells[7].text.strip(),
                        "去年同期": cells[8].text.strip(),
                        "与去年比": cells[9].text.strip(),
                    }
                    data.append(item)

                except StaleElementReferenceException:
                    logger.debug("Steel price: 检测到元素过期，重新获取表格数据...")
                    # 重新获取表格和行
                    table = driver.find_element(By.XPATH, '//table[contains(@class,"detailTab")]')
                    rows = table.find_elements(By.XPATH, './/tbody/tr[position()<=10]')
                    continue
                except Exception as e:
                    logger.debug(f"Steel price: 第 {idx} 行解析异常：{str(e)}")
                    continue

            logger.debug(f"成功抓取 Steel price 数据: {len(data)} 条记录")
            return data

        except TimeoutException:
            logger.error("Steel price: 页面加载超时，请检查网络连接或URL是否正确")
            return None
        except Exception as e:
            logger.error(f"Steel price: 爬取数据失败: {str(e)}")
            return None

    @log_execution_time
    @retry_on_timeout
    def crawl_shibor_rate(self, url):
        """
        爬取Shibor利率数据（优化版）
        """
        driver = self.get_driver(driver_type='daily')
        logger.debug(f"正在请求URL: {url}")

        try:
            # 设置页面加载超时
            driver.set_page_load_timeout(20)
            driver.get(url)

            # 使用显式等待，减少固定等待时间
            wait = WebDriverWait(driver, 10)
            table = wait.until(EC.presence_of_element_located((By.ID, 'shibor-tendays-show-data')))

            # 初始化结果数组
            result_list = []
            row_count = 0

            for row in table.find_elements(By.CSS_SELECTOR, "tr:has(td)"):
                if row_count >= 10:  # 修改为获取前10行数据
                    break  # 只取前10行数据

                cells = row.find_elements(By.TAG_NAME, "td")
                if len(cells) < 9:
                    continue

                # 解析数据
                current_record = {}
                current_record['日期'] = self.format_shibor_rate_date(cells[0].text.strip())
                terms = ['O/N', '1W', '2W', '1M', '3M', '6M', '9M', '1Y']

                for i, term in enumerate(terms):
                    current_record[term] = cells[i + 1].text.strip()

                result_list.append(current_record)
                row_count += 1

            logger.debug(f"成功抓取 Shibor 数据: {len(result_list)} 条记录")
            return result_list

        except TimeoutException:
            logger.error("Shibor: 页面加载超时，请检查网络连接或URL是否正确")
            return None
        except Exception as e:
            logger.error(f"Shibor: 数据抓取失败: {str(e)}")
            return None

    @log_execution_time
    @retry_on_timeout
    def crawl_lpr(self, url):
        """
        爬取LPR数据（优化版）
        """
        driver = self.get_driver(driver_type='daily')
        logger.debug(f"正在请求URL: {url}")

        try:
            # 设置页面加载超时
            driver.set_page_load_timeout(20)
            driver.get(url)

            # 使用显式等待，减少固定等待时间
            wait = WebDriverWait(driver, 10)
            table = wait.until(EC.presence_of_element_located((By.ID, 'lpr-ten-days-table')))

            # 提取关键数据
            rows = table.find_elements(By.CSS_SELECTOR, "tr")
            # 跳过表头行
            data_rows = rows[3:]

            # 初始化结果数组
            result_list = []
            row_index = 0

            for row in data_rows:
                if row_index >= 10:  # 修改为获取前10行数据
                    break

                cells = row.find_elements(By.TAG_NAME, "td")
                if len(cells) < 3:
                    continue

                date = self.format_lpr_date(cells[0].text.strip())
                term_1y = cells[1].text.strip()
                term_5y = cells[2].text.strip()

                current_record = {
                    "日期": date,
                    "1Y": term_1y,
                    "5Y": term_5y,
                    "PBOC_(6M-1Y)": 4.35,
                    "rowPBOC_(>5Y)": 4.9
                }
                result_list.append(current_record)
                row_index += 1

            logger.debug(f"成功抓取 LPR 数据: {len(result_list)} 条记录")
            return result_list

        except TimeoutException:
            logger.error("LPR: 页面加载超时，请检查网络连接或URL是否正确")
            return None
        except Exception as e:
            logger.error(f"LPR: 数据抓取失败: {str(e)}")
            return None

    @log_execution_time
    @retry_on_timeout
    def crawl_sofr(self, url):
        """
        爬取SOFR数据（优化版）
        """
        driver = self.get_driver(driver_type='daily')
        logger.debug(f"正在请求URL: {url}")

        try:
            # 设置页面加载超时
            driver.set_page_load_timeout(20)
            driver.get(url)

            # 使用显式等待，减少固定等待时间
            wait = WebDriverWait(driver, 10)
            table = wait.until(EC.presence_of_element_located((By.ID, 'pr_id_1-table')))

            # 获取所有数据行
            rows = table.find_elements(By.CSS_SELECTOR, "tr:has(td)")
            result_list = []

            # 处理前10行数据
            for row in rows[:10]:
                cells = row.find_elements(By.TAG_NAME, "td")

                # 确保列数足够
                if len(cells) < 7:
                    logger.debug(f"SOFR: 检测到不完整行，实际列数：{len(cells)}")
                    continue

                # 按顺序提取字段
                record = {
                    "日期": self.format_sofr_date(cells[0].text.strip()),
                    "Rate Type": 'SOFR',
                    "RATE(%)": cells[1].text.strip(),
                    "1ST PERCENTILE(%)": cells[2].text.strip(),
                    "25TH PERCENTILE(%)": cells[3].text.strip(),
                    "75TH PERCENTILE(%)": cells[4].text.strip(),
                    "99TH PERCENTILE(%)": cells[5].text.strip(),
                    "VOLUME ($Billions)": cells[6].text.strip()
                }
                result_list.append(record)

            logger.debug(f"成功抓取 SOFR 数据: {len(result_list)} 条记录")
            return result_list

        except TimeoutException:
            logger.error("SOFR: 页面加载超时，请检查网络连接或URL是否正确")
            return None
        except Exception as e:
            logger.error(f"SOFR: 数据抓取失败: {str(e)}")
            return None

    @log_execution_time
    @retry_on_timeout
    def crawl_ester(self, url):
        """
        爬取ESTER数据（优化版）
        """
        driver = self.get_driver(driver_type='daily')
        logger.debug(f"正在请求URL: {url}")

        try:
            # 设置页面加载超时
            driver.set_page_load_timeout(30)
            driver.get(url)

            # 使用显式等待，减少固定等待时间，增加超时时间
            wait = WebDriverWait(driver, 15)

            # 使用更精确的选择器
            tables = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "table.table-striped")))
            if not tables:
                logger.error("ESTER: 未找到目标表格")
                return None

            table = tables[0]  # 取第一个表格

            # 获取有效数据行（跳过表头）
            rows = table.find_elements(By.CSS_SELECTOR, "tr:has(td)")
            logger.debug(f"ESTER: 找到数据行数：{len(rows)}")

            result_list = []

            # 处理前10行数据
            for row in rows[:10]:
                cells = row.find_elements(By.TAG_NAME, "td")

                # 验证数据完整性
                if len(cells) != 2:
                    logger.debug(f"ESTER: 异常行数据，跳过。实际列数：{len(cells)}")
                    continue

                # 创建格式化记录
                record = {
                    "日期": self.format_ester_date(cells[0].get_attribute('textContent').strip()),
                    "value": cells[1].get_attribute('textContent').strip().replace(' %', '')
                }
                result_list.append(record)

            logger.debug(f"成功抓取 ESTER 数据: {len(result_list)} 条记录")
            return result_list

        except TimeoutException:
            logger.error("ESTER: 页面加载超时，请检查网络连接或URL是否正确")
            return None
        except Exception as e:
            logger.error(f"ESTER: 数据抓取异常: {str(e)}")
            return None

    @log_execution_time
    @retry_on_timeout
    def crawl_jpy_rate(self, url):
        """
        爬取JPY利率数据（优化版）
        """
        driver = self.get_driver(driver_type='daily')
        logger.debug(f"正在请求URL: {url}")

        try:
            # 设置页面加载超时
            driver.set_page_load_timeout(10)
            driver.get(url)

            # 使用显式等待，减少固定等待时间
            wait = WebDriverWait(driver, 10)

            # 使用更精确的选择器
            table = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "table.table[class='table ']")))

            # 获取有效数据行（跳过表头）
            rows = table.find_elements(By.CSS_SELECTOR, "tr:has(td)")
            result_list = []

            # 处理前10行数据
            for row in rows[:10]:
                cells = row.find_elements(By.TAG_NAME, "td")

                # 验证数据完整性
                if len(cells) != 2:
                    logger.debug(f"JPY rate: 异常行数据，跳过。实际列数：{len(cells)}")
                    continue

                # 创建格式化记录
                record = {
                    "日期": self.format_jpy_rate_date(cells[0].get_attribute('textContent').strip()),
                    "value": cells[1].get_attribute('textContent').strip().replace(' %', '')
                }
                result_list.append(record)

            logger.debug(f"成功抓取 JPY rate 数据: {len(result_list)} 条记录")
            return result_list

        except TimeoutException:
            logger.error("JPY rate: 页面加载超时，请检查网络连接或URL是否正确")
            return None
        except Exception as e:
            logger.error(f"JPY rate: 数据抓取异常: {str(e)}")
            return None

    @log_execution_time
    @retry_on_timeout
    def crawl_us_interest_rate(self, url):
        """
        爬取美国利率数据
        """
        driver = self.get_driver(driver_type='monthly')
        logger.debug(f"正在请求URL: {url}")

        try:
            # 设置页面加载超时
            driver.set_page_load_timeout(10)
            driver.get(url)

            # 使用显式等待，减少固定等待时间
            wait = WebDriverWait(driver, 10)
            table = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "table.table-model")))

            # 获取有效数据行（跳过表头）
            rows = table.find_elements(By.CSS_SELECTOR, "tr:has(td)")
            result_list = []

            # 处理前两行数据
            for row in rows[:2]:
                cells = row.find_elements(By.TAG_NAME, "td")

                # 验证数据完整性
                if len(cells) != 4:
                    logger.debug(f"US Interest Rate: 异常行数据，跳过。实际列数：{len(cells)}")
                    continue

                # 创建格式化记录
                record = {
                    "日期": cells[0].text.strip(),
                    "前值": cells[1].text.strip(),
                    "现值": cells[2].text.strip(),
                    "发布日期": self.format_us_interest_rate_date(cells[3].text.strip()),
                }
                result_list.append(record)

            logger.debug(f"成功抓取 US Interest Rate 数据: {len(result_list)} 条记录")
            return result_list

        except TimeoutException:
            logger.error("US Interest Rate: 页面加载超时，请检查网络连接或URL是否正确")
            return None
        except Exception as e:
            logger.error(f"US Interest Rate: 数据抓取异常: {str(e)}")
            return None

    @log_execution_time
    @retry_on_timeout
    def crawl_import_export(self, url):
        """
        爬取进出口贸易数据
        """
        driver = self.get_driver(driver_type='monthly')
        logger.debug(f"正在请求URL: {url}")

        try:
            # 设置页面加载超时
            driver.set_page_load_timeout(10)
            driver.get(url)

            # 使用显式等待，减少固定等待时间
            wait = WebDriverWait(driver, 10)
            table = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "table.table-model")))

            # 获取有效数据行（跳过表头）
            rows = table.find_elements(By.CSS_SELECTOR, "tr:has(td)")
            result_list = []

            # 处理前两行数据
            for row in rows[:2]:
                cells = row.find_elements(By.TAG_NAME, "td")

                # 验证数据完整性
                if len(cells) != 11:
                    logger.debug(f"Import Export: 异常行数据，跳过。实际列数：{len(cells)}")
                    continue

                # 创建格式化记录
                record = {
                    "日期": cells[0].text.strip(),
                    "当月出口额金额": cells[1].text.strip(),
                    "当月出口额同比增长": cells[2].text.strip(),
                    "当月出口额环比增长": cells[3].text.strip(),
                    "当月进口额金额": cells[4].text.strip(),
                    "当月进口额同比增长": cells[5].text.strip(),
                    "当月进口额环比增长": cells[6].text.strip(),
                    "累计出口额金额": cells[7].text.strip(),
                    "累计出口额同比增长": cells[8].text.strip(),
                    "累计进口额金额": cells[9].text.strip(),
                    "累计进口额同比增长": cells[10].text.strip(),
                }
                result_list.append(record)

            logger.debug(f"成功抓取 Import and Export 数据: {len(result_list)} 条记录")
            return result_list

        except TimeoutException:
            logger.error("Import Export: 页面加载超时，请检查网络连接或URL是否正确")
            return None
        except Exception as e:
            logger.error(f"Import Export: 数据抓取异常: {str(e)}")
            return None

    @log_execution_time
    @retry_on_timeout
    def crawl_money_supply(self, url):
        """
        爬取货币供应数据
        """
        driver = self.get_driver(driver_type='monthly')
        logger.debug(f"正在请求URL: {url}")

        try:
            # 设置页面加载超时
            driver.set_page_load_timeout(10)
            driver.get(url)

            # 使用显式等待，减少固定等待时间
            wait = WebDriverWait(driver, 10)
            table = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "table.table-model")))

            # 获取有效数据行（跳过表头）
            rows = table.find_elements(By.CSS_SELECTOR, "tr:has(td)")
            result_list = []

            # 处理前两行数据
            for row in rows[:2]:
                cells = row.find_elements(By.TAG_NAME, "td")

                # 验证数据完整性
                if len(cells) != 10:
                    logger.debug(f"Money Supply: 异常行数据，跳过。实际列数：{len(cells)}")
                    continue

                # 创建格式化记录
                record = {
                    "日期": cells[0].text.strip(),
                    "M2数量": cells[1].text.strip(),
                    "M2同比增长": cells[2].text.strip(),
                    "M2环比增长": cells[3].text.strip(),
                    "M1数量": cells[4].text.strip(),
                    "M1同比增长": cells[5].text.strip(),
                    "M1环比增长": cells[6].text.strip(),
                    "M0数量": cells[7].text.strip(),
                    "M0同比增长": cells[8].text.strip(),
                    "M0环比增长": cells[9].text.strip(),
                }
                result_list.append(record)

            logger.debug(f"成功抓取 Money Supply 数据: {len(result_list)} 条记录")
            return result_list

        except TimeoutException:
            logger.error("Money Supply: 页面加载超时，请检查网络连接或URL是否正确")
            return None
        except Exception as e:
            logger.error(f"Money Supply: 数据抓取异常: {str(e)}")
            return None

    @log_execution_time
    @retry_on_timeout
    def crawl_ppi(self, url):
        """
        爬取ppi数据
        """
        driver = self.get_driver(driver_type='monthly')
        logger.debug(f"正在请求URL: {url}")

        try:
            # 设置页面加载超时
            driver.set_page_load_timeout(10)
            driver.get(url)

            # 使用显式等待，减少固定等待时间
            wait = WebDriverWait(driver, 10)
            table = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "table.table-model")))

            # 获取有效数据行（跳过表头）
            rows = table.find_elements(By.CSS_SELECTOR, "tr:has(td)")
            result_list = []

            # 处理前两行数据
            for row in rows[:2]:
                cells = row.find_elements(By.TAG_NAME, "td")

                # 验证数据完整性
                if len(cells) != 4:
                    logger.debug(f"PPI: 异常行数据，跳过。实际列数：{len(cells)}")
                    continue

                # 创建格式化记录
                record = {
                    "日期": cells[0].text.strip(),
                    "当月": cells[1].text.strip(),
                    "当月同比增长": cells[2].text.strip(),
                    "累计": cells[3].text.strip(),
                }
                result_list.append(record)

            logger.debug(f"成功抓取 PPI 数据: {len(result_list)} 条记录")
            return result_list

        except TimeoutException:
            logger.error("PPI: 页面加载超时，请检查网络连接或URL是否正确")
            return None
        except Exception as e:
            logger.error(f"PPI: 数据抓取异常: {str(e)}")
            return None

    @log_execution_time
    @retry_on_timeout
    def crawl_cpi(self, url):
        """
        爬取cpi数据
        """
        driver = self.get_driver(driver_type='monthly')
        logger.debug(f"正在请求URL: {url}")

        try:
            # 设置页面加载超时
            driver.set_page_load_timeout(10)
            driver.get(url)

            # 使用显式等待，减少固定等待时间
            wait = WebDriverWait(driver, 10)
            table = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "table.table-model")))

            # 获取有效数据行（跳过表头）
            rows = table.find_elements(By.CSS_SELECTOR, "tr:has(td)")
            result_list = []

            # 处理前两行数据
            for row in rows[:2]:
                cells = row.find_elements(By.TAG_NAME, "td")

                # 验证数据完整性
                if len(cells) != 13:
                    logger.debug(f"CPI: 异常行数据，跳过。实际列数：{len(cells)}")
                    continue

                # 创建格式化记录
                record = {
                    "日期": cells[0].text.strip(),
                    "全国当月": cells[1].text.strip(),
                    "全国同比增长": cells[2].text.strip(),
                    "全国环比增长": cells[3].text.strip(),
                    "全国累计": cells[4].text.strip(),
                    "城市当月": cells[5].text.strip(),
                    "城市同比增长": cells[6].text.strip(),
                    "城市环比增长": cells[7].text.strip(),
                    "城市累计": cells[8].text.strip(),
                    "农村当月": cells[9].text.strip(),
                    "农村同比增长": cells[10].text.strip(),
                    "农村环比增长": cells[11].text.strip(),
                    "农村累计": cells[12].text.strip(),
                }
                result_list.append(record)

            logger.debug(f"成功抓取 CPI 数据: {len(result_list)} 条记录")
            return result_list

        except TimeoutException:
            logger.error("CPI: 页面加载超时，请检查网络连接或URL是否正确")
            return None
        except Exception as e:
            logger.error(f"CPI: 数据抓取异常: {str(e)}")
            return None

    @log_execution_time
    @retry_on_timeout
    def crawl_pmi(self, url):
        """
        爬取pmi数据
        """
        driver = self.get_driver(driver_type='monthly')
        logger.debug(f"正在请求URL: {url}")

        try:
            # 设置页面加载超时
            driver.set_page_load_timeout(10)
            driver.get(url)

            # 使用显式等待，减少固定等待时间
            wait = WebDriverWait(driver, 10)
            table = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "table.table-model")))

            # 获取有效数据行（跳过表头）
            rows = table.find_elements(By.CSS_SELECTOR, "tr:has(td)")
            result_list = []

            # 处理前两行数据
            for row in rows[:2]:
                cells = row.find_elements(By.TAG_NAME, "td")

                # 验证数据完整性
                if len(cells) != 5:
                    logger.debug(f"PMI: 异常行数据，跳过。实际列数：{len(cells)}")
                    continue

                # 创建格式化记录
                record = {
                    "日期": cells[0].text.strip(),
                    "制造业指数": cells[1].text.strip(),
                    "制造业同比增长": cells[2].text.strip(),
                    "非制造业指数": cells[3].text.strip(),
                    "非制造业同比增长": cells[4].text.strip(),
                }
                result_list.append(record)

            logger.debug(f"成功抓取 PMI 数据: {len(result_list)} 条记录")
            return result_list

        except TimeoutException:
            logger.error("PMI: 页面加载超时，请检查网络连接或URL是否正确")
            return None
        except Exception as e:
            logger.error(f"PMI: 数据抓取异常: {str(e)}")
            return None

    @log_execution_time
    @retry_on_timeout
    def crawl_new_bank_loan_addition(self, url):
        """
        爬取 中国 新增信贷数据
        """
        driver = self.get_driver(driver_type='monthly')
        logger.debug(f"正在请求URL: {url}")

        try:
            # 设置页面加载超时
            driver.set_page_load_timeout(10)
            driver.get(url)

            # 使用显式等待，减少固定等待时间
            wait = WebDriverWait(driver, 10)
            table = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "table.table-model")))

            # 获取有效数据行（跳过表头）
            rows = table.find_elements(By.CSS_SELECTOR, "tr:has(td)")
            result_list = []

            # 处理前两行数据
            for row in rows[:2]:
                cells = row.find_elements(By.TAG_NAME, "td")

                # 验证数据完整性
                if len(cells) != 6:
                    logger.debug(f"New Bank Loan: 异常行数据，跳过。实际列数：{len(cells)}")
                    continue

                # 创建格式化记录 - 修复字段名称，避免重复的"同比增长"
                record = {
                    "日期": cells[0].text.strip(),
                    "当月": cells[1].text.strip(),
                    "同比增长": cells[2].text.strip(),
                    "环比增长": cells[3].text.strip(),
                    "累计": cells[4].text.strip(),
                    "累计同比增长": cells[5].text.strip(),  # 修改为"累计同比增长"以区分
                }
                result_list.append(record)

            logger.debug(f"成功抓取 New Bank Loan Addition 数据: {len(result_list)} 条记录")
            return result_list

        except TimeoutException:
            logger.error("New Bank Loan: 页面加载超时，请检查网络连接或URL是否正确")
            return None
        except Exception as e:
            logger.error(f"New Bank Loan: 数据抓取异常: {str(e)}")
            return None

if __name__ == "__main__":
    def main():
        """主函数"""
        import argparse

        parser = argparse.ArgumentParser(description='市场数据爬取工具')
        parser.add_argument('--debug', action='store_true', help='启用调试日志')
        args = parser.parse_args()

        # 设置日志级别
        setup_logging(debug=args.debug)

        if args.debug:
            logger.info("启用调试模式，将显示详细日志")
        else:
            logger.info("使用标准日志级别。使用 --debug 参数可查看详细日志")

        print("==================================================")
        print("市场数据爬取工具")
        print("==================================================")

        analyzer = MarketDataAnalyzer()

        try:
            logger.info("开始更新市场数据...")
            analyzer.update_excel()
        except KeyboardInterrupt:
            logger.info("检测到用户中断，正在关闭资源...")
        except Exception as e:
            logger.error(f"程序执行出错: {str(e)}")
            analyzer.close_driver()
            analyzer.close_driver(driver_type='exchange_rate')  # 关闭汇率数据WebDriver
            analyzer.close_driver(driver_type='daily')         # 关闭日频数据WebDriver
            analyzer.close_driver(driver_type='monthly')       # 关闭月度数据WebDriver


        print("\n程序运行完成")

    main()
